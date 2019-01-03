from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook

# #打开一个excel文件,返回指定表
# def openExcel(pathname) -> None:
# 	wb = load_workbook(pathname)

# #根据表名获得excel文件中的某个表
# def getSheet(tableName):
# 	sheet = wb.get_sheet_by_name('Sheet1')		

#读取指定行数据
def read_row(sheet, row, column) -> list:
	src_row_data = []
	for cell in list(sheet.rows)[row][:column_index_from_string(column):]:
		src_row_data.append(str(cell.value).rstrip('\n'))
	#print(src_row_data)
	
	return src_row_data

#读取表中所有有效行数据
def read_all_rows(sheet):
	Signal_Datas = []
	for i in range(2, sheet.max_row):
		rowData = read_row(sheet, i, 'T')
		if rowData[0] != "None":
			Signal_Datas.append(rowData)

	return Signal_Datas

# 信号列表处理，将列表中的所有int转为str
def SignalDatas_handling(srcData:list) -> list:
	retDataList = []
	for i in range(len(srcData)):
		retDataList.append([str(j) for j in srcData[i]])

	return retDataList





#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping(sheet, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			ChannalMapping[sheet['W' + str(i + 3)].value] = sheet['V' + str(i + 3)].value.lower()
	elif dir == 1:
		for i in range(6):
			ChannalMapping[sheet['V' + str(i + 3)].value.lower()] = sheet['W' + str(i + 3)].value
	return ChannalMapping

#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping_pandas(dataFrame, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('N') - 1]] = dataFrame.values[i +2][column_index_from_string('M') - 1].lower()
			ChannalMapping[dataFrame.values[i + 2][1]] = dataFrame.values[i +2][0].lower()
	elif dir == 1:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('M') - 1]] = dataFrame.values[i +2][column_index_from_string('N') - 1]
			ChannalMapping[dataFrame.values[i + 2][0]] = dataFrame.values[i +2][1]
	return ChannalMapping




