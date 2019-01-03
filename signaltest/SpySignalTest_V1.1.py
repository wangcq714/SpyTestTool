import sys
import os
import platform

import pandas

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename

from report.testreport import *
from modules.readExcel import *
from modules.writefile import *
from config.config import *
from config import g_var


#生成SpyCCode.c函数
def SpyCCodemain():
	#调用选择文件夹窗口
	pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
	if pathname != ():
		#打开一个excel文件
		wb = load_workbook(pathname)
		#获取指定的表
		sheet = wb['Sheet2']
		#print(sheet.max_row)


		#读取所有的有效数据
		g_var.SignalDatas = read_all_rows(sheet)
		#print(g_var.SignalDatas)
		#print(len(g_var.SignalDatas))

		g_var.Num_Signal = len(g_var.SignalDatas)
		#print(g_var.Num_Signal)

		#获取通道映射
		g_var.ChannalMapping = getChannalMapping(sheet, 0)
		#print(g_var.ChannalMapping)
		g_var.ChannalMappingConvert = getChannalMapping(sheet, 1)
		#print(g_var.ChannalMappingConvert)

		#创建SpyCCode.c文件
		createSpyCCode(pathname)

		print('-'*20 + 'END' + '-'*20)

	else:
		print("没有选择路由表")
		exit()

#生成SpyCCode.c函数
def SpyCCodemain_pandas():
	#调用选择文件夹窗口
	pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
	if pathname != ():
		# 读取指定列有效数据
		dataFrame = pandas.read_excel(pathname, sheet_name="Sheet2", header=None, na_values="", usecols="A:T")
		# print(dataFrame)
		# 将DataFrame格式数据转为列表
		g_var.SignalDatas = dataFrame.fillna("None").values[2:].tolist()
		# print(g_var.SignalDatas)
		# 将列表中的所有int转为str
		g_var.SignalDatas = SignalDatas_handling(g_var.SignalDatas)
		# print(g_var.SignalDatas)
		g_var.Num_Signal = len(g_var.SignalDatas)
		# print(g_var.Num_Signal)

		# 读取指定列数据
		dataFrame = pandas.read_excel(pathname, sheet_name="Sheet2", header=None, na_values="", usecols="V:X")
		# print(dataFrame)
		# 获取通道映射
		g_var.ChannalMapping = getChannalMapping_pandas(dataFrame, 0)
		# print(g_var.ChannalMapping)
		g_var.ChannalMappingConvert = getChannalMapping_pandas(dataFrame, 1)
		# print(g_var.ChannalMappingConvert)

		# 创建SpyCCode.c文件
		createSpyCCode(pathname)


		print('-'*20 + 'END' + '-'*20)

	else:
		print("没有选择路由表")
		exit()

#生成测试报告函数
def reportmain():
	#获取测试log数据路径名
	pathname = askopenfilename(filetypes = [("TXT",".txt")])
	if pathname != ():
		#获取路由表路径名
		SignalTablepathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if SignalTablepathname != '':
			#打开一个excel文件
			wb = load_workbook(SignalTablepathname)
			#获取指定的表
			sheet = wb['Sheet2']
			#print(sheet.max_row)
			#读取所有的有效数据
			g_var.SignalDatas = read_all_rows(sheet)

			g_var.Num_Signal = len(g_var.SignalDatas)
			#设置表名
			SheetTitle = "SignalTestReport"			
			#获取测试报告表头列表
			TableHeaderList = build_TableHeaderList(sheet)
			#print(TableHeaderList)
			#读取SpyCCode生成的log日志
			DataList = readtestlog(pathname)
			#print(DataList)
			#创建测试报告列表
			TestReportList = create_Signal_testreportlist(TableHeaderList, DataList)
			#print(TestReportList)
			#设置保存生成报告的路径
			PathName = SignalTablepathname[:pathname.rfind('/')+1] + 'SignalTestReport.xlsx'
			#将测试报告数据写入excel
			write2excel(SheetTitle, TestReportList, PathName) 
			#设置测试报告Excel表格式
			setstyle(PathName, SheetTitle)

			#结束行
			print('-'*20 + "END" + '-'*20)

		else:
			print("没有选择路由表")
	else:
		print("没有选择data.txt文件")
		exit()

#生成测试报告函数
def reportmain_pandas():
	#获取测试log数据路径名
	pathname = askopenfilename(filetypes = [("TXT",".txt")])
	if pathname != ():
		#获取路由表路径名
		SignalTablepathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if SignalTablepathname != '':
			# 读取指定列有效数据
			dataFrame = pandas.read_excel(SignalTablepathname, sheet_name="Sheet2", header=None, na_values="", usecols="A:T")
			# print(dataFrame)
			# 将DataFrame格式数据转为列表
			g_var.SignalDatas = dataFrame.fillna("None").values[2:].tolist()
			# print(g_var.SignalDatas)
			# 将列表中的所有int转为str
			g_var.SignalDatas = SignalDatas_handling(g_var.SignalDatas)
			# print(g_var.SignalDatas)
			g_var.Num_Signal = len(g_var.SignalDatas)
			# print(g_var.Num_Signal)



			#设置表名
			SheetTitle = "SignalTestReport"			
			#获取测试报告表头列表
			TableHeaderList = build_TableHeaderList_pandas(dataFrame.fillna("None").values[0:2].tolist())
			#print(TableHeaderList)
			#读取SpyCCode生成的log日志
			DataList = readtestlog(pathname)
			#print(DataList)
			#创建测试报告列表
			TestReportList = create_Signal_testreportlist(TableHeaderList, DataList)
			#print(TestReportList)
			#设置保存生成报告的路径
			PathName = SignalTablepathname[:pathname.rfind('/')+1] + 'SignalTestReport.xlsx'
			#将测试报告数据写入excel
			write2excel(SheetTitle, TestReportList, PathName) 
			#设置测试报告Excel表格式
			setstyle(PathName, SheetTitle)

			#结束行
			print('-'*20 + "END" + '-'*20)

		else:
			print("没有选择路由表")
	else:
		print("没有选择data.txt文件")
		exit()

def main():
	#判断生成文件类型  1."Test":生成SpyCCode.c文件  2."Report":生成测试报告
	if ExecuteParameter == True:
		try:
			runPattern = sys.argv[1]
		except (IndexError, ):
			print("请给程序传参：python3 文件名.py Test or python3 文件名.py Report")
			exit()

		if sys.argv[1] == "Test":
			# SpyCCodemain()
			SpyCCodemain_pandas()
		elif sys.argv[1] == "Report":
			# reportmain()
			reportmain_pandas()
		else:
			print("RunPattern setup error!!!")
			exit()
	elif ExecuteParameter == False:
		if RunPattern[0] == "Test":
			# SpyCCodemain()
			SpyCCodemain_pandas()
		elif RunPattern[0] == "Report":
			# reportmain()
			reportmain_pandas()
		else:
			print("RunPattern setup error!!!")
			exit()

if __name__ == '__main__':
	main()

