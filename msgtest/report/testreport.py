from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font,colors,Border,Side,PatternFill,Alignment
from msgtest.modules.readExcel import *
# from msgtest.config.config import *
from msgtest.config import gv

#获取TestReportTableHeader
def get_TestReportTableHeader(sheet, num:int) -> list:
	TestReportTableHeaderList = []
	for i in range(num):
		TestReportTableHeaderList.append(read_line(sheet, i, 'L'))
	return TestReportTableHeaderList

#获取测试报告表头列表
def build_TableHeaderList(sheet) -> list:

	TableHeaderList = get_TestReportTableHeader(sheet, 2)
	TableHeaderList[0].append("TestResult")
	TableHeaderList[0].append("DelayTime")

	return TableHeaderList

#获取测试报告表头列表
def build_TableHeaderList_pandas(srcList) -> list:

	TableHeaderList = srcList
	TableHeaderList[0].append("TestResult")
	TableHeaderList[0].append("DelayTime")

	return TableHeaderList

def build_TableHeaderList_Direct() -> list:
	TableHeaderList = [['源报文ID', 'ECU', '周期', 'dlc', 'DTC码', '源网段', '目标网段', 'None', 'None', 'None', 'None', 'None', 'TestResult', 'DelayTime'], \
						['None', 'None', 'None', 'None', 'None', 'None', 'HybridCAN', 'PTCAN', 'EPCAN', 'BDCAN', 'CHCAN', 'DCAN']]
	return TableHeaderList


#读取测试MSGlog
def readtestlog(pathname: str) -> list:
	DataList = []
	with open(pathname) as Rfile:
		for tmp in Rfile:
			DataList.append(tmp)

	return DataList

#创建MSG测试列表
def createmsgtestreportlist(TableHeader: list, DataList: list) -> list:
	msgTestReportList = []
	msgTestReportList.extend(TableHeader)
	for i in range(len(gv.MessageData)):
		tmplist = []
		for tmp in gv.MessageData[i]:
			tmplist.append(tmp)
		if DataList[i].find("success") > 0:	
			tmplist.append("Success")
			tmplist.append(DataList[i][(DataList[i].find("DelayTime: ") + len("DelayTime: ")):])
			msgTestReportList.append(tmplist)
		elif DataList[i].find("fail") > 0:
			tmplist.append("Fail")
			msgTestReportList.append(tmplist)

	return msgTestReportList


#按行写入excel
def write2excel(SheetTitle: str, MSGTestReportList: list,  PathName: str) -> None:
	wb = Workbook()

	sheet=wb.worksheets[0]

	sheet.title = SheetTitle
	for tmp in MSGTestReportList:
		sheet.append(tmp)
	for i in range(sheet.max_row):
		for j in range(column_index_from_string('M')):
			if sheet[get_column_letter(j+1) + str(i + 1)].value == "None":
				sheet[get_column_letter(j+1) + str(i + 1)].value = ''

	#font = Font(color = colors.RED)
	#sheet["A1"].font = font
	wb.save(PathName)



#设置单元格格式
def setstyle(PathName:str, SheetTitle:str) -> None:
	# 默认可读写，若有需要可以指定write_only和read_only为True
	wb = load_workbook(PathName)
	sheet = wb[SheetTitle]
	for i in range(sheet.max_row):
		for j in range(column_index_from_string('N')):
			sheet[get_column_letter(j+1) + str(i + 1)].alignment = Alignment(horizontal = 'center', vertical = 'center')
			sheet[get_column_letter(j+1) + str(i + 1)].font = Font(name = 'Times New Roman')
			if i < 2:
				sheet[get_column_letter(j+1) + str(i + 1)].font = Font(name = '楷体', size = 12, bold = True)
		if sheet['M' + str(i + 1)].value == "Success":
			sheet['M' + str(i + 1)].fill = styles.fills.GradientFill(stop = ["00FF00", "00FF00"])
			sheet[get_column_letter(j) + str(i + 1)].font = Font(size = 12, bold = True)
		elif sheet['M' + str(i + 1)].value == "Fail":
			sheet['M' + str(i + 1)].fill = styles.fills.GradientFill(stop = ["FF0000", "FF0000"])
			sheet[get_column_letter(j+1) + str(i + 1)].font = Font(size = 12, bold = True)

	#设置行高
	for i in range(sheet.max_row):
		sheet.row_dimensions[i + 1].height = 18


	#设置列宽
	sheet.column_dimensions['B'].width = 20
	for i in range(column_index_from_string('G'),column_index_from_string('M')):
		sheet.column_dimensions[get_column_letter(i+1)].width = 10
	sheet.column_dimensions['M'].width = 15

	#合并单元格
	for i in range(column_index_from_string('F')):
		sheet.merge_cells(get_column_letter(i+1)+'1:'+get_column_letter(i+1)+'2')
	sheet.merge_cells("G1:L1")
	sheet.merge_cells("M1:M2")
	sheet.merge_cells("N1:N2")

	wb.save(PathName)














