from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook

#打开一个excel文件,返回指定表
def openExcel(pathname) -> None:
	wb = load_workbook(pathname)

#根据表名获得excel文件中的某个表
def getSheet(tableName):
	sheet = wb.get_sheet_by_name('Sheet1')		

#读取指定行
def read_line(sheet, row:int, tocolumn:str) -> list:
	src_row_data = []
	for cell in list(sheet.rows)[row][:column_index_from_string(tocolumn):]:
		src_row_data.append(str(cell.value).rstrip('\n'))
	return src_row_data

#读取所有行数据
def read_all_line(sheet) -> list:
	MessageData = []
	for i in range(2, sheet.max_row):
		#sheet['A' + str(i)].value为<class "str">类型，与"None"和None都不可比较
		#if sheet['A' + str(i)].value != None: 
			#print(sheet['A' + str(i)].value)
			#print(type(sheet['A' + str(i)].value))
		rowlist = read_line(sheet, i, 'L')
		if rowlist[column_index_from_string('A') - 1] != "None":
			MessageData.append(read_line(sheet, i, 'L'))
		i += 1
	return MessageData

#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping(sheet, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			ChannalMapping[sheet['N' + str(i + 3)].value] = sheet['M' + str(i + 3)].value.lower()
	elif dir == 1:
		for i in range(6):
			ChannalMapping[sheet['M' + str(i + 3)].value.lower()] = sheet['N' + str(i + 3)].value
	return ChannalMapping

#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping_pandas(dataFrame, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('N') - 1]] = dataFrame.values[i +2][column_index_from_string('M') - 1].lower()
			ChannalMapping[dataFrame.values[i + 2][1]] = dataFrame.values[i +2][0].lower()
	elif dir == 1:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('M') - 1]] = dataFrame.values[i +2][column_index_from_string('N') - 1]
			ChannalMapping[dataFrame.values[i + 2][0].lower()] = dataFrame.values[i +2][1]
	return ChannalMapping

# 将以行列表为单元的Excel数据（二维列表）指定列的None转为NA
def column_None2NA(srcData:list, column:str):
	for i in range(len(srcData)):
		if srcData[i][column_index_from_string(column) - 1] == "None":
			srcData[i][column_index_from_string(column) - 1] = "NA"
	# print(srcData)




