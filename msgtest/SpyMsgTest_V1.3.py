import sys
import os
import platform

import pandas

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename

from report.testreport import *
from modules.readExcel import *
from modules.writefile import *
from config.config import *
from config import gv


#生成SpyCCode.c函数
def SpyCCodemain():
	#调用选择文件夹窗口
	pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
	if pathname != ():
		#打开一个excel文件
		wb = load_workbook(pathname)
		#获取指定的表
		sheet = wb['Sheet1']
		#print(sheet.max_row)


		#读取所有的有效数据
		gv.MessageData = read_all_line(sheet)
		print(gv.MessageData)
		#print(len(gv.MessageData))

		#获取通道映射
		gv.ChannalMapping = getChannalMapping(sheet, 0)
		#print(gv.ChannalMapping)
		gv.ChannalMappingConvert = getChannalMapping(sheet, 1)
		#print(gv.ChannalMappingConvert)

		#创建SpyCCode.c文件
		createSpyCCode(pathname)

		print('-'*20 + 'END' + '-'*20)

	else:
		print("没有选择路由表")
		exit()

#生成SpyCCode.c函数
def SpyCCodemain_pandas():
	#调用选择文件夹窗口
	pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
	if pathname != ():
		# 读取指定列有效数据
		dataFrame = pandas.read_excel(pathname, sheet_name="Sheet1", header=None, na_values="", usecols="A:L")
		# print(dataFrame)
		gv.MessageData = dataFrame.fillna("None").values[2:].tolist()
		# print(gv.MessageData)
		# 将指定列中的None转为NA（为了重用之前版本的代码打个补丁）
		column_None2NA(gv.MessageData, 'C')
		# print(gv.MessageData)

		# 读取指定列数据
		dataFrame = pandas.read_excel(pathname, sheet_name="Sheet1", header=None, na_values="", usecols="M:N")
		# print(dataFrame)
		# 获取通道映射
		gv.ChannalMapping = getChannalMapping_pandas(dataFrame, 0)
		# print(gv.ChannalMapping)
		gv.ChannalMappingConvert = getChannalMapping_pandas(dataFrame, 1)
		# print(gv.ChannalMappingConvert)

		# 创建SpyCCode.c文件
		createSpyCCode(pathname)

		print('-'*20 + 'END' + '-'*20)

	else:
		print("没有选择路由表")
		exit()

#生成测试报告函数
def reportmain():
	#获取测试log数据路径名
	pathname = askopenfilename(filetypes = [("TXT",".txt")])
	if pathname != ():
		#获取路由表路径名
		RouterTablepathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if RouterTablepathname != '':
			#打开一个excel文件
			wb = load_workbook(RouterTablepathname)
			#获取指定的表
			sheet = wb['Sheet1']
			#print(sheet.max_row)
			#读取所有的有效数据
			gv.MessageData = read_all_line(sheet)

			#设置表名
			SheetTitle = "MessageRouteTestReport"			
			#获取测试报告表头列表
			TableHeaderList = build_TableHeaderList(sheet)
			#读取SpyCCode生成的log日志
			DataList = readtestlog(pathname)
			#创建测试报告列表
			TestReportList = createmsgtestreportlist(TableHeaderList, DataList)
			#print(testreportlist)
			#设置保存生成报告的路径
			PathName = RouterTablepathname[:pathname.rfind('/')+1] + 'MessageRouteReport.xlsx'
			#将测试报告数据写入excel
			write2excel(SheetTitle, TestReportList, PathName) 
			#设置测试报告Excel表格式
			setstyle(PathName, SheetTitle)

			#结束行
			print('-'*20 + "END" + '-'*20)

		else:
			print("没有选择路由表")
	else:
		print("没有选择data.txt文件")
		exit()

#生成测试报告函数
def reportmain_pandas():
	#获取测试log数据路径名
	pathname = askopenfilename(filetypes = [("TXT", ".txt")])
	if pathname != ():
		#获取路由表路径名
		RouterTablepathname = askopenfilename(filetypes = [("Excel", ".xlsx")])
		if RouterTablepathname != '':
			# 读取所有的有效数据
			dataFrame = pandas.read_excel(RouterTablepathname, sheet_name="Sheet1", header=None, na_values="", usecols="A:L")
			# print(dataFrame)
			gv.MessageData = dataFrame.fillna("None").values[2:].tolist()
			# 将指定列中的None转为NA（为了重用之前的代码打个补丁）
			column_None2NA(gv.MessageData, 'C')
			# print(gv.MessageData)

			#设置表名
			SheetTitle = "MessageRouteTestReport"			
			#获取测试报告表头列表
			TableHeaderList = build_TableHeaderList_pandas(dataFrame.fillna("None").values[0:2].tolist())
			# print(TableHeaderList)
			#读取SpyCCode生成的log日志
			DataList = readtestlog(pathname)
			#创建测试报告列表
			TestReportList = createmsgtestreportlist(TableHeaderList, DataList)
			#print(testreportlist)
			#设置保存生成报告的路径
			PathName = RouterTablepathname[:pathname.rfind('/')+1] + 'MessageRouteReport.xlsx'
			#将测试报告数据写入excel
			write2excel(SheetTitle, TestReportList, PathName) 
			#设置测试报告Excel表格式
			setstyle(PathName, SheetTitle)

			#结束行
			print('-'*20 + "END" + '-'*20)

		else:
			print("没有选择路由表")
	else:
		print("没有选择data.txt文件")
		exit()

def main():
	#判断生成文件类型  1."Default" + "NA"默认生成SpyCCode.c文件  2."Report" + "NA"生成测试报告 3."Clone" + "*CAN"生成克隆测试SpyCCode.c
	if ExecuteParameter == True:
		try:
			runPattern = sys.argv[1]
		except (IndexError, ):
			print("请给程序传参：python3 文件名.py Default NA or python3 文件名.py Clone XXX or python3 文件名.py Report NA")
			exit()

		if sys.argv[1] == "Default" or sys.argv[1] == "Clone":
			# SpyCCodemain()
			SpyCCodemain_pandas()
		elif sys.argv[1] == "Report":
			# reportmain()
			reportmain_pandas()
		else:
			print("RunPattern setup error!!!")
			exit()

	elif ExecuteParameter == False:
		if RunPattern[0] == "Default" or RunPattern[0] == "Clone":
			# SpyCCodemain()
			SpyCCodemain_pandas()
		elif RunPattern[0] == "Report":
			# reportmain()
			reportmain_pandas()
		else:
			print("RunPattern setup error!!!")
			exit()


if __name__ == '__main__':
	main()
