# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
# from openpyxl import Workbook

# #打开一个excel文件,返回指定表
# def openExcel(pathname) -> None:
# 	wb = load_workbook(pathname)

# #根据表名获得excel文件中的某个表
# def getSheet(tableName):
# 	sheet = wb.get_sheet_by_name('Sheet1')		

#读取指定行数据
def read_row(sheet, row, column) -> list:
	src_column_data = []
	for cell in list(sheet.rows)[row][:column_index_from_string(column):]:
		src_column_data.append(str(cell.value).rstrip('\n'))
	#print(src_row_data)
	
	return src_column_data

#读取表中所有有效行数据
def read_all_rows(sheet):
	DTC_Datas = []
	for i in range(2, sheet.max_row):
		tmpData = read_row(sheet, i, 'J')
		if tmpData[0] == "None" or tmpData[0] == "NA":
			continue
		else:
			DTC_Datas.append(tmpData)

	return DTC_Datas

# 获取表中节点丢失有效行数据
def DTC_Datas_handling(srcData:list) -> list:
	DTC_Datas = []
	for i in range(len(srcData)):
		if srcData[i][0] == "None" or srcData[i][0] == "NA":
			continue
		else:
			srcData[i] = [str(j) for j in srcData[i]]
			DTC_Datas.append(srcData[i])

	return DTC_Datas




#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping(sheet, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			ChannalMapping[sheet['N' + str(i + 3)].value] = sheet['M' + str(i + 3)].value.lower()
	elif dir == 1:
		for i in range(6):
			ChannalMapping[sheet['M' + str(i + 3)].value.lower()] = sheet['N' + str(i + 3)].value
	return ChannalMapping

#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping_pandas(dataFrame, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('N') - 1]] = dataFrame.values[i +2][column_index_from_string('M') - 1].lower()
			ChannalMapping[dataFrame.values[i + 2][1]] = dataFrame.values[i +2][0].lower()
	elif dir == 1:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('M') - 1]] = dataFrame.values[i +2][column_index_from_string('N') - 1]
			ChannalMapping[dataFrame.values[i + 2][0]] = dataFrame.values[i +2][1]
	return ChannalMapping

#读取参数设置
def get_para_set(sheet) -> tuple:
	para_set = {}
	for i in range(3):
		para_set[sheet['M' + str(i + 12)].value.strip(' ')] = sheet['N' + str(i + 12)].value.strip(' ')

	return (para_set["DiagCH"], para_set["Diag_Req_ID"], para_set["Diag_Res_ID"])

#读取参数设置
def get_para_set_pandas(dataFrame) -> tuple:
	para_set = {}
	for i in range(3):
		para_set[dataFrame.values[i + 12 - 1][0].strip(' ')] = dataFrame.values[i + 12 - 1][1].strip(' ')

	return (para_set["DiagCH"], para_set["Diag_Req_ID"], para_set["Diag_Res_ID"])


